import sys, csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

# Ambil argumen path CSV
csv_path = sys.argv[1]
output_path = "/tmp/output.xlsx"

wb = Workbook()

# --- Sheet 1: Data ---
ws_data = wb.active
ws_data.title = "Data"

# Styling header (kuning, tebal, hitam)
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_font = Font(bold=True, color="000000")

col_widths = {}

# Load CSV ke sheet Data
with open(csv_path, newline="", encoding="utf-8") as f:
    reader = csv.reader(f)
    for row_index, row in enumerate(reader, start=1):
        for col_index, value in enumerate(row, start=1):
            cell = ws_data.cell(row=row_index, column=col_index, value=value)

            # Hitung max panjang text
            if value:
                col_widths[col_index] = max(col_widths.get(col_index, 0), len(str(value)))

            # Styling header (baris pertama)
            if row_index == 1:
                cell.fill = header_fill
                cell.font = header_font
            else:
                # Konversi NOMER STORE jadi angka
                header_value = ws_data.cell(row=1, column=col_index).value
                if header_value and "NOMER STORE" in header_value.upper():
                    val = str(value).strip()
                    if val.isdigit():
                        cell.value = int(val)
                    else:
                        try:
                            cell.value = float(val.replace(",", ""))
                        except:
                            pass

# Set lebar kolom otomatis
for col_index, width in col_widths.items():
    col_letter = get_column_letter(col_index)
    ws_data.column_dimensions[col_letter].width = width + 2

# --- Sheet 2: Summary ---
ws_sum = wb.create_sheet("Summary")

# Cari kolom T. TRX RITEL & T. REV RITEL
headers = [cell.value for cell in ws_data[1]]
trx_col = headers.index("T. TRX RITEL") + 1 if "T. TRX RITEL" in headers else None
rev_col = headers.index("T. REV RITEL") + 1 if "T. REV RITEL" in headers else None
last_row = ws_data.max_row

ws_sum["A1"] = "Metric"
ws_sum["B1"] = "Formula"
ws_sum["C1"] = "Result"

metrics = []
if trx_col:
    trx_range = f"{get_column_letter(trx_col)}2:{get_column_letter(trx_col)}{last_row}"
    metrics.extend([
        ("Total TRX Ritel", f"=SUM({trx_range})"),
        ("Rata-rata TRX Ritel", f"=AVERAGE({trx_range})"),
        ("Max TRX Ritel", f"=MAX({trx_range})"),
        ("Min TRX Ritel", f"=MIN({trx_range})"),
    ])

if rev_col:
    rev_range = f"{get_column_letter(rev_col)}2:{get_column_letter(rev_col)}{last_row}"
    metrics.extend([
        ("Total REV Ritel", f"=SUM({rev_range})"),
        ("Rata-rata REV Ritel", f"=AVERAGE({rev_range})"),
        ("Max REV Ritel", f"=MAX({rev_range})"),
        ("Min REV Ritel", f"=MIN({rev_range})"),
    ])

row = 2
for name, formula in metrics:
    ws_sum[f"A{row}"] = name
    ws_sum[f"B{row}"] = formula
    ws_sum[f"C{row}"] = formula  # Excel akan evaluate
    row += 1

# Simpan file
wb.save(output_path)

print("✅ Conversion done, file saved to", output_path)




# import sys, csv
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter

# csv_path = sys.argv[1]
# output_path = sys.argv[2]

# print(">>> CSV input:", csv_path)
# print(">>> Excel output:", output_path)

# wb = Workbook()
# ws_data = wb.active
# ws_data.title = "Data"

# # Load CSV
# with open(csv_path, newline="", encoding="utf-8") as f:
#     reader = csv.reader(f)
#     for row in reader:
#         ws_data.append(row)

# # Pastikan ada header
# if ws_data.max_row < 1:
#     raise Exception("CSV kosong atau tidak ada header")

# headers = [cell.value for cell in ws_data[1]]
# print(">>> Headers:", headers)

# # Cari kolom
# trx_col = headers.index("T. TRX RITEL") + 1
# rev_col = headers.index("T. REV RITEL") + 1
# last_row = ws_data.max_row

# # Buat summary
# ws_sum = wb.create_sheet("Summary")
# ws_sum["A1"] = "Metric"
# ws_sum["B1"] = "Formula"
# ws_sum["C1"] = "Result"

# trx_range = f"{get_column_letter(trx_col)}2:{get_column_letter(trx_col)}{last_row}"
# rev_range = f"{get_column_letter(rev_col)}2:{get_column_letter(rev_col)}{last_row}"

# metrics = [
#     ("Total TRX Ritel", f"=SUM({trx_range})"),
#     ("Rata-rata TRX Ritel", f"=AVERAGE({trx_range})"),
#     ("Max TRX Ritel", f"=MAX({trx_range})"),
#     ("Min TRX Ritel", f"=MIN({trx_range})"),
#     ("Total REV Ritel", f"=SUM({rev_range})"),
#     ("Rata-rata REV Ritel", f"=AVERAGE({rev_range})"),
#     ("Max REV Ritel", f"=MAX({rev_range})"),
#     ("Min REV Ritel", f"=MIN({rev_range})"),
# ]

# row = 2
# for name, formula in metrics:
#     ws_sum[f"A{row}"] = name
#     ws_sum[f"B{row}"] = formula
#     ws_sum[f"C{row}"] = formula
#     row += 1

# wb.save(output_path)
# print(">>> Done. Saved:", output_path)