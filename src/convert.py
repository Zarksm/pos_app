# import sys
# import pandas as pd
# from openpyxl import load_workbook
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import PatternFill

# if len(sys.argv) < 3:
#     print("Usage: convert.py input.csv output.xlsx")
#     sys.exit(1)

# input_file = sys.argv[1]
# output_file = sys.argv[2]

# # 1. Baca CSV dengan parsing angka
# df = pd.read_csv(input_file, thousands=",", decimal=",")

# # 2. Simpan ke Excel
# df.to_excel(output_file, index=False)

# # 3. Load lagi untuk nambah formula & styling
# wb = load_workbook(output_file)
# ws = wb.active

# # ====== STYLE HEADER ======
# header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# for cell in ws[1]:
#     cell.fill = header_fill

# # ====== FORMULA TOTAL REV ======
# last_col = ws.max_column + 1
# ws.cell(row=1, column=last_col, value="Total Rev")

# for row in range(2, ws.max_row + 1):
#     formula = f"=SUM(G{row},I{row},M{row},P{row},S{row},V{row})"
#     ws.cell(row=row, column=last_col, value=formula)

# # ====== AUTO WIDTH ======
# for col in ws.columns:
#     max_length = 0
#     col_letter = get_column_letter(col[0].column)
#     for cell in col:
#         try:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         except:
#             pass
#     ws.column_dimensions[col_letter].width = max_length + 2

# wb.save(output_file)
# print("Conversion done!")



import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font

if len(sys.argv) < 3:
    print("Usage: convert.py input.csv output.xlsx")
    sys.exit(1)

input_file = sys.argv[1]
output_file = sys.argv[2]

# 1. Baca CSV dengan parsing angka
df = pd.read_csv(input_file, thousands=",", decimal=",")

# 2. Simpan ke Excel
df.to_excel(output_file, index=False)

# 3. Load lagi untuk nambah formula & styling
wb = load_workbook(output_file)
ws = wb.active

# ====== STYLE HEADER ======
header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
header_font = Font(bold=True)

ws.row_dimensions[1].height = 30  # set tinggi row pertama

for cell in ws[1]:
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.font = header_font

# ====== FORMULA TOTAL REV ======
last_col = ws.max_column + 1
ws.cell(row=1, column=last_col, value="Total Rev")
ws.cell(row=1, column=last_col).fill = header_fill
ws.cell(row=1, column=last_col).alignment = header_alignment
ws.cell(row=1, column=last_col).font = header_font

for row in range(2, ws.max_row + 1):
    formula = f"=SUM(G{row},I{row},M{row},P{row},S{row},V{row})"
    ws.cell(row=row, column=last_col, value=formula)

# ====== AUTO WIDTH ======
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(output_file)
print("Conversion done!")
